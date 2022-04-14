from flask import abort, send_file
from flask import current_app as app
from flask import flash, redirect, render_template, request, url_for
from flask_user import current_user
from opencve.commands import info
from opencve.constants import PRODUCT_SEPARATOR
from opencve.controllers.main import main, welcome
from opencve.controllers.reports import ReportController
from opencve.extensions import db
from opencve.models.changes import Change
from opencve.models.categories import Category
from opencve.models.users import User
from opencve.models.cve import Cve
from opencve.models.events import Event
from sqlalchemy import and_, or_, func
from sqlalchemy.dialects.postgresql import array
from sqlalchemy.orm import aliased, joinedload
from flask_login import current_user, login_required
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import datetime
from opencve.forms import ActivitiesViewForm
from opencve.utils import (
    convert_cpes,
    get_cpe_list_from_specific_product,
    get_cwes_details,
    CustomHtmlHTML,
)


from celery.utils.log import get_task_logger


logger = get_task_logger(__name__)


@welcome.route("/welcome")
def index():
    if not app.config.get("DISPLAY_WELCOME", False):
        abort(404)
    return render_template("index.html")


@welcome.route("/terms")
def terms():
    if not app.config.get("DISPLAY_TERMS", False):
        abort(404)
    return render_template("terms.html")


@main.route("/", methods=["GET", "POST"])
def home():
    # Allow customization of the homepage
    if not current_user.is_authenticated:
        if app.config.get("DISPLAY_WELCOME", False):
            return redirect(url_for("welcome.index"))
        return redirect(url_for("main.cves"))

    # Form used to customize the activities view
    activities_view_form = ActivitiesViewForm(
        obj=current_user,
        view=current_user.settings["activities_view"],
    )

    if request.method == "POST":
        form_name = request.form["form-name"]
        if form_name == "activities_view_form" and activities_view_form.validate():
            new_settings = {
                **current_user.settings,
                "activities_view": activities_view_form.view.data,
            }
            current_user.settings = new_settings
            db.session.commit()

            flash("Your settings has been updated.", "success")
            return redirect(url_for("main.home"))

    # For every category the user follows, we also add the products and vendors
    # Fetch the user subscriptions
    vendors = [v.name for v in current_user.vendors]
    for c in current_user.categories:
        vendors.extend([f"{v.name}" for v in c.vendors])
        vendors.extend(
            [f"{p.vendor.name}{PRODUCT_SEPARATOR}{p.name}" for p in c.products]
        )

    # Handle the page parameter
    page = request.args.get("page", type=int, default=1)
    page = 1 if page < 1 else page
    per_page = app.config["ACTIVITIES_PER_PAGE"]

    # Only display the 5 last reports
    reports = ReportController.list_items({"user_id": current_user.id})[:5]

    # Build the query to fetch the last changes
    query = (
        Change.query.options(joinedload("cve"))
        .options(joinedload("events"))
        .filter(Change.cve_id == Cve.id)
        .filter(Change.events.any())
    )

    # Filter by subscriptions
    if current_user.settings["activities_view"] == "subscriptions":
        vendors = [v.name for v in current_user.vendors]
        for p in current_user.products:
            cpes = get_cpe_list_from_specific_product(p)
            vendors.extend(cpes)
        for c in current_user.categories:
            vendors.extend([f"{v.name}" for v in c.vendors])
            vendors.extend([f"{v.name}" for v in c.vendors])
            for p in c.products:
                cpes = get_cpe_list_from_specific_product(p)
                vendors.extend(cpes)
        if not vendors:
            vendors = [None]
        query = query.filter(Cve.vendors.has_any(array(vendors)))

    # List the paginated changes
    changes = (
        query.order_by(Change.created_at.desc())
        .limit(per_page)
        .offset((page - 1) * per_page)
        .all()
    )

    return render_template(
        "home.html",
        changes=changes,
        reports=reports,
        page=page,
        activities_view_form=activities_view_form,
    )


@main.route("/generateUserReport")
@login_required
def generateUserReport():
    """Generate a report for the specified user"""
    # UPLOAD_FOLDER = '/app/venv/lib/python3.7/site-packages/opencve/data/'
    # Create a xlsx file with the user name
    # The file is named after the username and the datetime
    file_name = (
        str(current_user.username)
        + "_"
        + str(datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S"))
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Product Report"

    # Print the products of the user and the number of CVEs associated to them
    date = datetime.datetime.now() - datetime.timedelta(days=30)
    ws["A1"] = "Product"
    ws["B1"] = "Vendor"
    ws["C1"] = "Number of CVEs since" + str(date.strftime("%Y-%m-%d"))
    ws["D1"] = "CVSS2 Score mean"
    ws["E1"] = "CVSS3 Score mean"
    ws["F1"] = "Number of critical CVEs (above 7.5)"
    i = 2
    total = 0
    cves = []

    for product in current_user.products:
        cpes = get_cpe_list_from_specific_product(product)
        ws["A" + str(i)] = product.name
        ws["B" + str(i)] = product.vendor.name
        count = 0
        cveQuery = Cve.query.filter(
            and_(
                Cve.vendors.has_any(array(cpes)) if product else None,
                Cve.updated_at >= date,
            )
        )

        cves += cveQuery
        count = cveQuery.count()
        total += count
        ws["C" + str(i)] = count

        cvss2AVG = (
            cveQuery.filter(Cve.cvss2 != None)
            .with_entities(func.avg(Cve.cvss2))
            .scalar()
        )
        cvss3AVG = (
            cveQuery.filter(Cve.cvss3 != None)
            .with_entities(func.avg(Cve.cvss3))
            .scalar()
        )

        cvss2AVG = round(cvss2AVG, 2) if cvss2AVG else None
        cvss3AVG = round(cvss3AVG, 2) if cvss3AVG else None

        # TODO : The colors are not working properly because cvss2 and cvss3 are NoneType for some reason
        ws["D" + str(i)] = cvss2AVG if cvss2AVG else None  # "N/A"
        # if cvss2AVG >= 7.5:
        #     ws['D'+ str(i)].fill = PatternFill(fgColor="00FF0000", fill_type="solid")
        # elif cvss2AVG >= 5.0:
        #     ws['D'+ str(i)].fill = PatternFill(fgColor="00FF6600", fill_type="solid")
        # else:
        #     ws['D'+str(i)].fill = PatternFill(fgColor="00FFFF00", fill_type="solid")

        ws["E" + str(i)] = cvss3AVG if cvss3AVG else None  # "N/A"
        # if cvss3AVG >= 7.5:
        #     ws['E'+str(i)].fill = PatternFill(fgColor="00FF0000", fill_type="solid")
        # elif cvss3AVG >= 5.0:
        #     ws['E'+str(i)].fill = PatternFill(fgColor="00FF6600", fill_type="solid")
        # else:
        #     ws['E'+str(i)].fill = PatternFill(fgColor="00FFFF00", fill_type="solid")

        criticalCount = cveQuery.filter(or_(Cve.cvss2 >= 7.5, Cve.cvss3 >= 7.5)).count()

        ws["F" + str(i)] = criticalCount
        if count and criticalCount / count >= 0.75:
            ws["F" + str(i)].fill = PatternFill(fgColor="00FF0000", fill_type="solid")
        elif count and criticalCount / count >= 0.25:
            ws["F" + str(i)].fill = PatternFill(fgColor="00FF6600", fill_type="solid")
        elif count and criticalCount / count > 0.0:
            ws["F" + str(i)].fill = PatternFill(fgColor="00FFFF00", fill_type="solid")

        i += 1

    ws["A" + str(i)] = "Totals and averages"
    ws["C" + str(i)] = "=SUM(C2:C" + str(i - 1) + ")"
    ws["D" + str(i)] = "=AVERAGE(D2:D" + str(i - 1) + ")"
    ws["E" + str(i)] = "=AVERAGE(E2:E" + str(i - 1) + ")"
    ws["F" + str(i)] = "=SUM(F2:F" + str(i - 1) + ")"

    ws = wb.create_sheet("Vendors Report")
    ws["A1"] = "Vendor"
    ws["B1"] = "Number of CVEs since" + str(date.strftime("%Y-%m-%d"))
    ws["C1"] = "CVSS2 Score mean"
    ws["D1"] = "CVSS3 Score mean"
    ws["E1"] = "Number of critical CVEs (above 7.5)"
    i = 2
    total = 0
    for vendor in current_user.vendors:
        ws["A" + str(i)] = vendor.name
        count = 0
        cveQuery = Cve.query.filter(
            and_(
                or_(
                    Cve.vendors.contains([vendor.name])
                    if vendor
                    else None,  # For the moment, the count is also based on the vendor
                    # Cve.vendors.contains([product.vendor.name+'$PRODUCT$'+product.name]) if product else None,
                ),
                Cve.updated_at >= date,
            )
        )

        cvss2AVG = (
            cveQuery.filter(Cve.cvss2 != None)
            .with_entities(func.avg(Cve.cvss2))
            .scalar()
        )
        cvss3AVG = (
            cveQuery.filter(Cve.cvss3 != None)
            .with_entities(func.avg(Cve.cvss3))
            .scalar()
        )

        cvss2AVG = round(cvss2AVG, 2) if cvss2AVG else None
        cvss3AVG = round(cvss3AVG, 2) if cvss3AVG else None

        # TODO : The colors are not working properly because cvss2 and cvss3 are NoneType for some reason
        ws["C" + str(i)] = cvss2AVG if cvss2AVG else None  # "N/A"
        # if cvss2AVG >= 7.5:
        #     ws['C'+ str(i)].fill = PatternFill(fgColor="00FF0000", fill_type="solid")
        # elif cvss2AVG >= 5.0:
        #     ws['C'+ str(i)].fill = PatternFill(fgColor="00FF6600", fill_type="solid")
        # else:
        #     ws['C'+str(i)].fill = PatternFill(fgColor="00FFFF00", fill_type="solid")
        ws["D" + str(i)] = cvss3AVG if cvss3AVG else None  # "N/A"
        # if cvss3AVG >= 7.5:
        #     ws['D'+str(i)].fill = PatternFill(fgColor="00FF0000", fill_type="solid")
        # elif cvss3AVG >= 5.0:
        #     ws['D'+str(i)].fill = PatternFill(fgColor="00FF6600", fill_type="solid")
        # else:
        #     ws['D'+str(i)].fill = PatternFill(fgColor="00FFFF00", fill_type="solid")

        criticalCount = cveQuery.filter(or_(Cve.cvss2 >= 7.5, Cve.cvss3 >= 7.5)).count()

        cves += cveQuery
        count = cveQuery.count()
        ws["C" + str(i)] = count
        ws["E" + str(i)] = criticalCount
        if count and criticalCount / count >= 0.75:
            ws["E" + str(i)].fill = PatternFill(fgColor="00FF0000", fill_type="solid")
        elif count and criticalCount / count >= 0.25:
            ws["E" + str(i)].fill = PatternFill(fgColor="00FF6600", fill_type="solid")
        elif count and criticalCount / count > 0.0:
            ws["E" + str(i)].fill = PatternFill(fgColor="00FFFF00", fill_type="solid")

        i += 1
    ws["A" + str(i)] = "Totals and averages"
    ws["B" + str(i)] = "=SUM(B2:B" + str(i - 1) + ")"
    ws["C" + str(i)] = "=AVERAGE(C2:C" + str(i - 1) + ")"
    ws["D" + str(i)] = "=AVERAGE(D2:D" + str(i - 1) + ")"
    ws["E" + str(i)] = "=SUM(E2:E" + str(i - 1) + ")"

    # Print all the CVEs associated to the user
    ws = wb.create_sheet("CVEs Report")
    ws["A1"] = "CVE ID"
    ws["B1"] = "Last update"
    ws["C1"] = "Creation date"
    ws["D1"] = "Vendor"
    ws["E1"] = "Product"
    ws["F1"] = "Description"
    ws["G1"] = "CWE"
    ws["H1"] = "CVSS2"
    ws["I1"] = "CVSS3"
    ws["J1"] = "Mitre Link"
    ws["K1"] = "NVD Link"

    i = 2
    cves.sort(key=lambda x: x.updated_at, reverse=True)
    for cve in cves:
        ws["A" + str(i)] = cve.cve_id
        ws["B" + str(i)] = cve.updated_at.strftime("%Y-%m-%d")
        ws["C" + str(i)] = cve.created_at.strftime("%Y-%m-%d")
        CVEvendors = []
        CVEproducts = []
        tab = convert_cpes(cve.json["configurations"])
        for vendor in tab:
            CVEvendors.append(vendor)
            for product in tab[vendor]:
                CVEproducts.append(product)
        ws["D" + str(i)] = str(CVEvendors)
        ws["E" + str(i)] = str(CVEproducts)
        ws["F" + str(i)] = cve.summary

        cwes = get_cwes_details(
            cve.json["cve"]["problemtype"]["problemtype_data"][0]["description"]
        )
        ws["G" + str(i)] = str(cwes)

        ws["H" + str(i)] = cve.cvss2
        # if cve.cvss2 >= 7.5:
        #     ws['H'+ str(i)].fill = PatternFill(fgColor="00FF0000", fill_type="solid")
        # elif cve.cvss2 >= 5.0:
        #     ws['H'+ str(i)].fill = PatternFill(fgColor="00FF6600", fill_type="solid")
        # else:
        #     ws['H'+str(i)].fill = PatternFill(fgColor="00FFFF00", fill_type="solid")
        ws["I" + str(i)] = cve.cvss3
        # if cve.cvss3 >= 7.5:
        #     ws['I'+ str(i)].fill = PatternFill(fgColor="00FF0000", fill_type="solid")
        # elif cve.cvss3 >= 5.0:
        #     ws['I'+ str(i)].fill = PatternFill(fgColor="00FF6600", fill_type="solid")
        # else:
        #     ws['I'+str(i)].fill = PatternFill(fgColor="00FFFF00", fill_type="solid")

        # ws['J'+str(i)] = "https://cve.mitre.org/cgi-bin/cvename.cgi?name="+cve.cve_id
        ws["J" + str(i)] = (
            '=HYPERLINK("https://cve.mitre.org/cgi-bin/cvename.cgi?name='
            + cve.cve_id
            + '")'
        )
        ws["J" + str(i)].style = "Hyperlink"
        # ws['K'+str(i)] = "https://nvd.nist.gov/vuln/detail/"+cve.cve_id
        ws["K" + str(i)] = (
            '=HYPERLINK("https://nvd.nist.gov/vuln/detail/' + cve.cve_id + '")'
        )
        ws["K" + str(i)].style = "Hyperlink"

        i += 1

    ws["A" + str(i)] = "Totals and averages"
    ws["H" + str(i)] = "=AVERAGE(H2:H" + str(i - 1) + ")"
    ws["I" + str(i)] = "=AVERAGE(I2:I" + str(i - 1) + ")"

    wb.save("/" + file_name + ".xlsx")
    return send_file(
        "/" + file_name + ".xlsx",
        as_attachment=True,
        attachment_filename=file_name + ".xlsx",
    )
