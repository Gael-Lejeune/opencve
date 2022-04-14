from asyncio.log import logger
from ctypes import sizeof
import sys
from difflib import get_close_matches
from pathlib import Path
from opencve.extensions import cel
import openpyxl
from openpyxl.styles import PatternFill
from flask import abort, send_file
from flask import current_app as app
from opencve.commands import create_category, ensure_config, error, info
from opencve.controllers.base import BaseController
from opencve.extensions import db
from opencve.models.categories import Category
from opencve.models.products import Product
from opencve.models.vendors import Vendor
from opencve.models.cve import Cve
from opencve.models.alerts import Alert
from opencve.utils import get_categories_letters
from sqlalchemy.exc import IntegrityError
import datetime
from openpyxl import Workbook
from flask_user import current_user
from sqlalchemy import and_, or_, func
from opencve.utils import convert_cpes, get_cwes_details, CustomHtmlHTML
from opencve.controllers.cves import CveController


class CategoryController(BaseController):
    model = Category
    order = Category.name.asc()
    per_page_param = "CATEGORYS_PER_PAGE"
    schema = {
        "letter": {"type": str},
        "search": {"type": str},
        "category_name": {"type": str},
    }

    @classmethod
    def build_query(cls, args):
        letter = args.get("letter")

        query = cls.model.query

        # Search by term
        if args.get("search"):
            search = args.get("search").lower().replace("%", "").replace("_", "")
            query = query.filter(cls.model.name.like("%{}%".format(search)))

        # Search by letter
        if letter:
            if letter not in get_categories_letters():
                abort(404)

            query = query.filter(cls.model.name.like("{}%".format(letter)))

        return query, {}


VENDORS = []
PRODUCTS = []


def dehumanize_filter(name):
    """Reverses opencve.context._humanize_filter"""
    return "_".join(map(lambda x: x.lower(), name.split(" ")))


def find_db_name(Model, name):
    """Uses difflib.get_close_matches to find the closest match to name in the Model specified
    Made to work only with Vendor and Product"""
    try:
        name = dehumanize_filter(name)
        human_names = VENDORS if Model is Vendor else PRODUCTS

        db_name = get_close_matches(name, human_names, n=1, cutoff=0.6)
        if len(db_name) > 0:
            db_name = db_name[0]
        else:
            raise LookupError
        return db_name
    except LookupError:
        return
    except Exception as e:
        error(e)
        return


def add_product(category, tag):
    """Add Product to category"""

    product = Product.query.filter_by(name=tag).first()
    if product in category.products or product is None:
        return
    else:
        category.products.append(product)
        try:
            db.session.commit()
            info(f"[ADD_PRODUCT] {product.name} added to category {category}")
            return
        except IntegrityError as e:
            error(e)
            return -1


def create_category(name):
    """Create a Category."""
    name = str(name).lower()
    if Category.query.filter_by(name=name).first():
        return -1

    category = Category(
        name=name,
    )
    db.session.add(category)
    try:
        db.session.commit()
    except IntegrityError as e:
        error(e)
        return -1


def edit_category_name(category, name):
    """Edit the specified Category name."""
    name = str(name).lower()
    if Category.query.filter_by(name=name).first():
        return -1
    else:
        category.name = name
        try:
            db.session.commit()
        except IntegrityError as e:
            error(e)
            return -1


def delete_category(category):
    """Delete the specified Category."""
    try:
        db.session.delete(category)
        db.session.commit()
    except IntegrityError as e:
        error(e)
        return -1


@cel.task(bind=True)
def import_from_excel(self, category_name, path_to_file):
    cel.app.app_context().push()
    """Read an xlsx file and expects from it to have columns named vendor, product, version and tag
    Those names shall be found in the first three rows"""
    # with app.app_context():
    category = Category.query.filter_by(name=category_name).first()
    xlsx_file = open(path_to_file, "rb")
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active
    data = []
    # vendor_col = None
    # product_col = None
    # version_col = None
    tag_col = None
    min_value_index = None
    max_value_index = sheet.max_row

    for row in range(1, 3):  # To be sure, we check the 3 firsts rows
        for col in range(0, sheet.max_column):
            if str(sheet[row][col].value).lower() == "tag":
                tag_col = col
                min_value_index = row + 1
                break

    # For every row, we check if the value is already in the data list
    if min_value_index is None:
        return -1
    for i in range(min_value_index, max_value_index):
        if tag_col != None:
            tag = str(sheet[i][tag_col].value)
            if ":" in tag:
                # info(f"[IMPORT_FROM_EXCEL] : in {tag}")
                tag = tag.split(":")[4]
                # info(f"[IMPORT_FROM_EXCEL] tag changed to {tag}")
        else:
            tag = None
        if tag not in data:
            data.append(tag)
    ite = 0
    for tag in data:
        ite += 1
        # info(f"[IMPORT_FROM_EXCEL] {ite}/{len(data)}")
        # info(f"[IMPORT_FROM_EXCEL] {tag}")
        add_product(category, tag)
    return


def generateCategoryReport(category, period):
    """Generate a report for the specified category"""
    # UPLOAD_FOLDER = '/app/venv/lib/python3.7/site-packages/opencve/data/'
    # Create a xlsx file with the category name
    # The file is named after the category and the datetime
    file_name = (
        str(category.name)
        + "_"
        + str(datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S"))
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Product Report"

    # Print the products of the category and the number of CVEs associated to them
    date = datetime.datetime.now() - datetime.timedelta(days=period)
    ws["A1"] = "Product"
    ws["B1"] = "Vendor"
    ws["C1"] = "Number of CVEs since" + str(date.strftime("%Y-%m-%d"))
    ws["D1"] = "CVSS2 Score mean"
    ws["E1"] = "CVSS3 Score mean"
    ws["F1"] = "Number of critical CVEs (CVSS2 or CVSS3 above 7.5)"
    i = 2
    cves = []

    for product in category.products:
        ws["A" + str(i)] = product.name
        ws["B" + str(i)] = product.vendor.name
        count = 0
        cveQuery = Cve.query.filter(
            and_(
                or_(
                    # Cve.vendors.contains([product.vendor.name]) if vendor else None, # For the moment, the count is also based on the vendor
                    Cve.vendors.contains(
                        [product.vendor.name + "$PRODUCT$" + product.name]
                    )
                    if product
                    else None,
                ),
                Cve.updated_at >= date,
            )
        )

        cves += cveQuery
        count = cveQuery.count()

        ws["C" + str(i)] = count

        cvss2AVG = (
            cveQuery.filter(Cve.cvss2 != None)
            .with_entities(func.avg(Cve.cvss2))
            .scalar()
        )
        cvss3AVG = (
            cveQuery.filter(Cve.cvss3 != None)
            .with_entities(func.avg(Cve.cvss3))
            .scalar()
        )

        cvss2AVG = round(cvss2AVG, 2) if cvss2AVG else None
        cvss3AVG = round(cvss3AVG, 2) if cvss3AVG else None

        # TODO : The colors are not working properly because cvss2 and cvss3 are NoneType for some reason
        ws["D" + str(i)] = cvss2AVG if cvss2AVG else None  # "N/A"
        # if cvss2AVG >= 7.5:
        #     ws['D'+ str(i)].fill = PatternFill(fgColor="00FF0000", fill_type="solid")
        # elif cvss2AVG >= 5.0:
        #     ws['D'+ str(i)].fill = PatternFill(fgColor="00FF6600", fill_type="solid")
        # else:
        #     ws['D'+str(i)].fill = PatternFill(fgColor="00FFFF00", fill_type="solid")

        ws["E" + str(i)] = cvss3AVG if cvss3AVG else None  # "N/A"
        # if cvss3AVG >= 7.5:
        #     ws['E'+str(i)].fill = PatternFill(fgColor="00FF0000", fill_type="solid")
        # elif cvss3AVG >= 5.0:
        #     ws['E'+str(i)].fill = PatternFill(fgColor="00FF6600", fill_type="solid")
        # else:
        #     ws['E'+str(i)].fill = PatternFill(fgColor="00FFFF00", fill_type="solid")

        criticalCount = cveQuery.filter(or_(Cve.cvss2 >= 7.5, Cve.cvss3 >= 7.5)).count()

        ws["F" + str(i)] = criticalCount
        if count and criticalCount / count >= 0.75:
            ws["F" + str(i)].fill = PatternFill(fgColor="00FF0000", fill_type="solid")
        elif count and criticalCount / count >= 0.25:
            ws["F" + str(i)].fill = PatternFill(fgColor="00FF6600", fill_type="solid")
        elif count and criticalCount / count > 0.0:
            ws["F" + str(i)].fill = PatternFill(fgColor="00FFFF00", fill_type="solid")

        i += 1

    ws["A" + str(i)] = "Totals and averages"
    ws["C" + str(i)] = "=SUM(C2:C" + str(i - 1) + ")"
    ws["D" + str(i)] = "=AVERAGE(D2:D" + str(i - 1) + ")"
    ws["E" + str(i)] = "=AVERAGE(E2:E" + str(i - 1) + ")"
    ws["F" + str(i)] = "=SUM(F2:F" + str(i - 1) + ")"

    ws = wb.create_sheet("Vendors Report")
    ws["A1"] = "Vendor"
    ws["B1"] = "Number of CVEs since" + str(date.strftime("%Y-%m-%d"))
    ws["C1"] = "CVSS2 Score mean"
    ws["D1"] = "CVSS3 Score mean"
    ws["E1"] = "Number of critical CVEs (CVSS2 or CVSS3 above 7.5)"
    i = 2

    for vendor in category.vendors:
        ws["A" + str(i)] = vendor.name
        count = 0
        cveQuery = Cve.query.filter(
            and_(
                or_(
                    Cve.vendors.contains([vendor.name])
                    if vendor
                    else None,  # For the moment, the count is also based on the vendor
                    # Cve.vendors.contains([product.vendor.name+'$PRODUCT$'+product.name]) if product else None,
                ),
                Cve.updated_at >= date,
            )
        )

        cvss2AVG = (
            cveQuery.filter(Cve.cvss2 != None)
            .with_entities(func.avg(Cve.cvss2))
            .scalar()
        )
        cvss3AVG = (
            cveQuery.filter(Cve.cvss3 != None)
            .with_entities(func.avg(Cve.cvss3))
            .scalar()
        )

        cvss2AVG = round(cvss2AVG, 2) if cvss2AVG else None
        cvss3AVG = round(cvss3AVG, 2) if cvss3AVG else None

        # TODO : The colors are not working properly because cvss2 and cvss3 are NoneType for some reason
        ws["C" + str(i)] = cvss2AVG if cvss2AVG else None  # "N/A"
        # if cvss2AVG >= 7.5:
        #     ws['C'+ str(i)].fill = PatternFill(fgColor="00FF0000", fill_type="solid")
        # elif cvss2AVG >= 5.0:
        #     ws['C'+ str(i)].fill = PatternFill(fgColor="00FF6600", fill_type="solid")
        # else:
        #     ws['C'+str(i)].fill = PatternFill(fgColor="00FFFF00", fill_type="solid")
        ws["D" + str(i)] = cvss3AVG if cvss3AVG else None  # "N/A"
        # if cvss3AVG >= 7.5:
        #     ws['D'+str(i)].fill = PatternFill(fgColor="00FF0000", fill_type="solid")
        # elif cvss3AVG >= 5.0:
        #     ws['D'+str(i)].fill = PatternFill(fgColor="00FF6600", fill_type="solid")
        # else:
        #     ws['D'+str(i)].fill = PatternFill(fgColor="00FFFF00", fill_type="solid")

        criticalCount = cveQuery.filter(or_(Cve.cvss2 >= 7.5, Cve.cvss3 >= 7.5)).count()

        cves += cveQuery
        count = cveQuery.count()
        ws["B" + str(i)] = count
        ws["E" + str(i)] = criticalCount
        if count and criticalCount / count >= 0.75:
            ws["E" + str(i)].fill = PatternFill(fgColor="00FF0000", fill_type="solid")
        elif count and criticalCount / count >= 0.25:
            ws["E" + str(i)].fill = PatternFill(fgColor="00FF6600", fill_type="solid")
        elif count and criticalCount / count > 0.0:
            ws["E" + str(i)].fill = PatternFill(fgColor="00FFFF00", fill_type="solid")

        i += 1
    ws["A" + str(i)] = "Totals and averages"
    ws["B" + str(i)] = "=SUM(B2:B" + str(i - 1) + ")"
    ws["C" + str(i)] = "=AVERAGE(C2:C" + str(i - 1) + ")"
    ws["D" + str(i)] = "=AVERAGE(D2:D" + str(i - 1) + ")"
    ws["E" + str(i)] = "=SUM(E2:E" + str(i - 1) + ")"

    # Print all the CVEs associated to the category
    ws = wb.create_sheet("CVEs Report")
    ws["A1"] = "CVE ID"
    ws["B1"] = "Last update"
    ws["C1"] = "Creation date"
    ws["D1"] = "Vendor"
    ws["E1"] = "Product"
    ws["F1"] = "Description"
    ws["G1"] = "CWE"
    ws["H1"] = "CVSS2"
    ws["I1"] = "CVSS3"
    ws["J1"] = "Mitre Link"
    ws["K1"] = "NVD Link"

    i = 2

    cves.sort(key=lambda x: x.updated_at, reverse=True)
    for cve in cves:
        ws["A" + str(i)] = cve.cve_id
        ws["B" + str(i)] = cve.updated_at.strftime("%Y-%m-%d")
        ws["C" + str(i)] = cve.created_at.strftime("%Y-%m-%d")
        CVEvendors = []
        CVEproducts = []
        tab = convert_cpes(cve.json["configurations"])

        for vendor in tab:
            CVEvendors.append(vendor)
            for product in tab[vendor]:
                CVEproducts.append(product)
        ws["D" + str(i)] = str(CVEvendors)
        ws["E" + str(i)] = str(CVEproducts)
        ws["F" + str(i)] = cve.summary

        cwes = get_cwes_details(
            cve.json["cve"]["problemtype"]["problemtype_data"][0]["description"]
        )
        ws["G" + str(i)] = str(cwes)

        ws["H" + str(i)] = cve.cvss2
        # if cve.cvss2 >= 7.5:
        #     ws['H'+ str(i)].fill = PatternFill(fgColor="00FF0000", fill_type="solid")
        # elif cve.cvss2 >= 5.0:
        #     ws['H'+ str(i)].fill = PatternFill(fgColor="00FF6600", fill_type="solid")
        # else:
        #     ws['H'+str(i)].fill = PatternFill(fgColor="00FFFF00", fill_type="solid")
        ws["I" + str(i)] = cve.cvss3
        # if cve.cvss3 >= 7.5:
        #     ws['I'+ str(i)].fill = PatternFill(fgColor="00FF0000", fill_type="solid")
        # elif cve.cvss3 >= 5.0:
        #     ws['I'+ str(i)].fill = PatternFill(fgColor="00FF6600", fill_type="solid")
        # else:
        #     ws['I'+str(i)].fill = PatternFill(fgColor="00FFFF00", fill_type="solid")

        # ws['J'+str(i)] = "https://cve.mitre.org/cgi-bin/cvename.cgi?name="+cve.cve_id
        ws["J" + str(i)] = (
            '=HYPERLINK("https://cve.mitre.org/cgi-bin/cvename.cgi?name='
            + cve.cve_id
            + '")'
        )
        ws["J" + str(i)].style = "Hyperlink"
        # ws['K'+str(i)] = "https://nvd.nist.gov/vuln/detail/"+cve.cve_id
        ws["K" + str(i)] = (
            '=HYPERLINK("https://nvd.nist.gov/vuln/detail/' + cve.cve_id + '")'
        )
        ws["K" + str(i)].style = "Hyperlink"
        i += 1

    ws["A" + str(i)] = "Totals and averages"
    ws["H" + str(i)] = "=AVERAGE(H2:H" + str(i - 1) + ")"
    ws["I" + str(i)] = "=AVERAGE(I2:I" + str(i - 1) + ")"

    wb.save("/tmp/shared/" + file_name + ".xlsx")
    return send_file(
        "/tmp/shared/" + file_name + ".xlsx",
        as_attachment=True,
        attachment_filename=file_name + ".xlsx",
    )
